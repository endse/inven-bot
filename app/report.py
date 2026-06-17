import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models import Product, Transaction

def generate_monthly_report(db: Session, year_month: str, excel_path: str = "GST 2026.xlsx") -> str:
    """
    Generates a monthly sheet in the Excel file.
    year_month format: YYYY-MM (e.g. '2026-06')
    Sheet name: MMM-YY (e.g. 'Jun-26')
    """
    # Parse target month
    target_dt = datetime.strptime(year_month, "%Y-%m")
    sheet_name = target_dt.strftime("%b-%y")  # 'Jun-26'
    purchase_col_name = f"{year_month} purchase"  # '2026-06 purchase'

    # Load existing workbook or create new one
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

    # If sheet already exists, remove it so we can overwrite it
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(title=sheet_name)

    # Styles matching the original
    tnr_font = Font(name="Times New Roman", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )

    # Define headers
    headers = [
        "S.no",
        "Name",
        "H.S.N no",
        "Rate",
        "Opening balance",
        purchase_col_name,
        "Sales",
        "balance stock"
    ]

    # Write headers
    ws.append(headers)
    for col_idx in range(1, 9):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = tnr_font
        cell.alignment = center_align

    # Set column widths matching original
    column_widths = {
        'A': 8,   # S.no
        'B': 35,  # Name
        'C': 15,  # HSN
        'D': 12,  # Rate
        'E': 18,  # Opening Balance
        'F': 18,  # Monthly Purchase
        'G': 12,  # Sales
        'H': 18   # Balance Stock
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Fetch all products ordered by name or by database id
    products = db.query(Product).order_by(Product.id).all()

    # Define date thresholds
    start_of_month = target_dt.date()
    # Next month's first day
    if target_dt.month == 12:
        start_of_next_month = datetime(target_dt.year + 1, 1, 1).date()
    else:
        start_of_next_month = datetime(target_dt.year, target_dt.month + 1, 1).date()

    for i, product in enumerate(products, start=1):
        # 1. Opening Balance = Sum of all purchases minus sales before the start of the month
        # purchases before start of month
        purchases_before = db.query(func.sum(Transaction.quantity)).filter(
            Transaction.product_id == product.id,
            Transaction.transaction_type == 'purchase',
            Transaction.transaction_date < start_of_month
        ).scalar() or 0.0

        sales_before = db.query(func.sum(Transaction.quantity)).filter(
            Transaction.product_id == product.id,
            Transaction.transaction_type == 'sale',
            Transaction.transaction_date < start_of_month
        ).scalar() or 0.0

        opening_balance = float(purchases_before - sales_before)

        # 2. Monthly Purchases
        monthly_purchases = db.query(func.sum(Transaction.quantity)).filter(
            Transaction.product_id == product.id,
            Transaction.transaction_type == 'purchase',
            Transaction.transaction_date >= start_of_month,
            Transaction.transaction_date < start_of_next_month
        ).scalar() or 0.0
        monthly_purchases = float(monthly_purchases)

        # 3. Monthly Sales
        monthly_sales = db.query(func.sum(Transaction.quantity)).filter(
            Transaction.product_id == product.id,
            Transaction.transaction_type == 'sale',
            Transaction.transaction_date >= start_of_month,
            Transaction.transaction_date < start_of_next_month
        ).scalar() or 0.0
        monthly_sales = float(monthly_sales)

        # 4. Balance Stock
        balance_stock = opening_balance + monthly_purchases - monthly_sales

        row_data = [
            i,
            product.name,
            product.hsn or "",
            float(product.last_rate) if product.last_rate else "",
            opening_balance if opening_balance != 0 else "",
            monthly_purchases if monthly_purchases != 0 else "",
            monthly_sales if monthly_sales != 0 else "",
            balance_stock if balance_stock != 0 else ""
        ]

        ws.append(row_data)

        # Style the row cells
        row_idx = i + 1
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = tnr_font
            cell.border = thin_border
            
            if col_idx == 1:  # S.no
                cell.alignment = center_align
            elif col_idx == 2:  # Name
                cell.alignment = left_align
            elif col_idx in [3, 4]:  # HSN, Rate
                cell.alignment = center_align
                if col_idx == 4 and cell.value != "":  # Rate formatting
                    # Format as rupee if we want, or standard decimal. Let's use number format
                    cell.number_format = '₹#,##0.00'
            else:  # Balances, Purchase, Sales
                cell.alignment = right_align

    wb.save(excel_path)
    return excel_path
