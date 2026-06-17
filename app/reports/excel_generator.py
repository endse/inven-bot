import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database.models import Product, Transaction

def generate_monthly_report(db: Session, year_month: str, excel_path: str = "GST 2026.xlsx") -> str:
    """
    Generates a monthly sheet in the Excel file.
    year_month format: YYYY-MM (e.g. '2026-06')
    """
    target_dt = datetime.strptime(year_month, "%Y-%m")
    sheet_name = target_dt.strftime("%b-%y")
    purchase_col_name = f"{year_month} purchase"

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(title=sheet_name)

    tnr_font = Font(name="Times New Roman", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3")
    )

    headers = [
        "S.no", "Name", "H.S.N no", "Rate", 
        "Opening balance", purchase_col_name, "Sales", "balance stock"
    ]

    ws.append(headers)
    for col_idx in range(1, 9):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = tnr_font
        cell.alignment = center_align

    column_widths = {'A': 8, 'B': 35, 'C': 15, 'D': 12, 'E': 18, 'F': 18, 'G': 12, 'H': 18}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    products = db.query(Product).order_by(Product.id).all()
    start_of_month = target_dt.date()
    
    if target_dt.month == 12:
        start_of_next_month = datetime(target_dt.year + 1, 1, 1).date()
    else:
        start_of_next_month = datetime(target_dt.year, target_dt.month + 1, 1).date()

    for i, product in enumerate(products, start=1):
        purchases_before = db.query(func.sum(Transaction.quantity)).filter(
            Transaction.product_id == product.id,
            Transaction.transaction_type == 'purchase',
            Transaction.transaction_date < start_of_month
        ).scalar() or 0.0

        sales_before = db.query(func.sum(Transaction.quantity)).filter(
            Transaction.product_id == product.id,
            Transaction.transaction_type == 'sale',
            Transaction.transaction_date < start_of_month
        ).scalar() or 0.0

        opening_balance = float(purchases_before - sales_before)

        monthly_purchases = float(db.query(func.sum(Transaction.quantity)).filter(
            Transaction.product_id == product.id,
            Transaction.transaction_type == 'purchase',
            Transaction.transaction_date >= start_of_month,
            Transaction.transaction_date < start_of_next_month
        ).scalar() or 0.0)

        monthly_sales = float(db.query(func.sum(Transaction.quantity)).filter(
            Transaction.product_id == product.id,
            Transaction.transaction_type == 'sale',
            Transaction.transaction_date >= start_of_month,
            Transaction.transaction_date < start_of_next_month
        ).scalar() or 0.0)

        balance_stock = opening_balance + monthly_purchases - monthly_sales

        row_data = [
            i, product.name, product.hsn or "", float(product.last_rate) if product.last_rate else "",
            opening_balance if opening_balance != 0 else "",
            monthly_purchases if monthly_purchases != 0 else "",
            monthly_sales if monthly_sales != 0 else "",
            balance_stock if balance_stock != 0 else ""
        ]

        ws.append(row_data)

        row_idx = i + 1
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = tnr_font
            cell.border = thin_border
            
            if col_idx == 1:
                cell.alignment = center_align
            elif col_idx == 2:
                cell.alignment = left_align
            elif col_idx in [3, 4]:
                cell.alignment = center_align
                if col_idx == 4 and cell.value != "":
                    cell.number_format = '₹#,##0.00'
            else:
                cell.alignment = right_align

    wb.save(excel_path)
    return excel_path

def generate_yearly_report(db: Session, year: str, excel_path: str = "GST 2026.xlsx") -> str:
    """
    Generates all 12 monthly sheets for the given year.
    year format: YYYY (e.g. '2026')
    """
    for month in range(1, 13):
        year_month = f"{year}-{month:02d}"
        generate_monthly_report(db, year_month, excel_path)
    return excel_path
